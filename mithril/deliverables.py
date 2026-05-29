"""Deliverable finalization: agent Markdown → real Office files.

The agent always writes Markdown (what models are best at; what the grader
ultimately reads back as text anyway). This module converts each requested
deliverable into a genuine `.docx`/`.xlsx` so the judge's reader
(`pandoc` for docx, `pandas.read_excel` for xlsx) sees well-structured
content — it is the exact inverse of the grader's extraction path.

It is robust to the agent mis-naming or mis-formatting: if a deliverable was
saved directly under its binary name as plain Markdown text, we detect that
(not a valid OOXML zip) and re-render it.
"""

from __future__ import annotations

import re
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook


def _is_ooxml(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        with zipfile.ZipFile(path) as z:
            return "[Content_Types].xml" in z.namelist()
    except (zipfile.BadZipFile, OSError):
        return False


def _stem_words(name: str) -> set[str]:
    return set(Path(name).stem.lower().replace("-", " ").replace("_", " ").split())


def _find_markdown_source(output_dir: Path, deliverable: str) -> Path | None:
    """Locate the Markdown the agent intended for this deliverable."""
    stem = Path(deliverable).stem
    exact = output_dir / f"{stem}.md"
    if exact.exists():
        return exact

    mds = [p for p in output_dir.rglob("*.md") if p.is_file()]
    want = _stem_words(deliverable)
    best, best_score = None, 0
    for p in mds:
        score = len(_stem_words(p.name) & want)
        if score > best_score:
            best, best_score = p, score
    if best is not None and best_score > 0:
        return best

    # Agent may have written the deliverable directly under its binary name but
    # as Markdown text (invalid OOXML). Treat that as the source.
    direct = output_dir / deliverable
    if direct.exists() and not _is_ooxml(direct):
        return direct

    # Single .md fallback.
    if len(mds) == 1:
        return mds[0]
    return None


def _render_docx(md_text: str, dest: Path) -> bool:
    with tempfile.NamedTemporaryFile("w", suffix=".md", delete=False, encoding="utf-8") as tmp:
        tmp.write(md_text)
        tmp_path = tmp.name
    try:
        r = subprocess.run(
            ["pandoc", tmp_path, "-f", "gfm", "-o", str(dest)],
            capture_output=True, text=True, timeout=120,
        )
        return r.returncode == 0 and dest.exists()
    except (OSError, subprocess.TimeoutExpired):
        return False
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def _parse_md_tables(md_text: str) -> list[tuple[str, list[list[str]]]]:
    """Return [(heading, rows)] for each Markdown table, paired with the nearest preceding heading."""
    tables, heading = [], "Sheet1"
    lines = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        h = re.match(r"^#{1,6}\s+(.*)", line)
        if h:
            heading = h.group(1).strip()[:31] or "Sheet"
        if line.startswith("|") and i + 1 < len(lines) and re.match(r"^\|?[\s:|-]+\|?$", lines[i + 1].strip()):
            rows, j = [], i
            while j < len(lines) and lines[j].strip().startswith("|"):
                cells = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                rows.append(cells)
                j += 1
            # Drop the separator row (row index 1).
            if len(rows) >= 2:
                rows = [rows[0]] + rows[2:]
            tables.append((heading, rows))
            i = j
            continue
        i += 1
    return tables


def _render_xlsx(md_text: str, dest: Path) -> bool:
    try:
        wb = Workbook()
        wb.remove(wb.active)
        tables = _parse_md_tables(md_text)
        if tables:
            used = set()
            for idx, (heading, rows) in enumerate(tables):
                title = re.sub(r"[\\/*?:\[\]]", " ", heading)[:31] or f"Sheet{idx+1}"
                base, n = title, 1
                while title.lower() in used:
                    n += 1
                    title = f"{base[:28]}_{n}"
                used.add(title.lower())
                ws = wb.create_sheet(title=title)
                for row in rows:
                    ws.append(row)
        else:
            ws = wb.create_sheet(title="Sheet1")
            for line in md_text.splitlines():
                ws.append([line])
        wb.save(dest)
        return dest.exists()
    except Exception:  # noqa: BLE001
        return False


def finalize_deliverables(output_dir: Path, expected: list[str]) -> dict:
    """Ensure every requested deliverable exists as a valid Office file. Returns a report."""
    output_dir = Path(output_dir)
    report = {}
    for deliverable in expected:
        dest = output_dir / deliverable
        ext = dest.suffix.lower()

        if _is_ooxml(dest):
            report[deliverable] = "kept (valid)"
            continue

        src = _find_markdown_source(output_dir, deliverable)
        if src is None:
            report[deliverable] = "MISSING (no source markdown)"
            continue
        md_text = src.read_text(encoding="utf-8", errors="replace")

        ok = False
        if ext == ".docx":
            ok = _render_docx(md_text, dest)
        elif ext == ".xlsx":
            ok = _render_xlsx(md_text, dest)
        else:
            # Unknown/extension-less: just copy the markdown content under the name.
            dest.write_text(md_text, encoding="utf-8")
            ok = True
        report[deliverable] = f"rendered from {src.name}" if ok else f"RENDER FAILED from {src.name}"
    return report
